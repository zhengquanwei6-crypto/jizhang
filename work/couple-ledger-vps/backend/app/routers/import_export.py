from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime, timedelta
from typing import Any, Literal

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response, StreamingResponse

from app.auth import get_current_user
from app.db import apply_balance_delta, get_db, local_today, new_id, row_to_dict, utcnow
from app.routers.transactions import _validate_scope_access, _validate_split_fields
from app.ws import manager

router = APIRouter(prefix="/api/data", tags=["data"])

ALIPAY_MARKERS = ("支付宝", "alipay", "交易号", "收/支")
WECHAT_MARKERS = ("微信支付", "wechat", "微信昵称", "当前状态")


def _detect_format(text: str) -> str:
    head = text[:2000].lower()
    if any(m.lower() in head for m in WECHAT_MARKERS):
        return "wechat"
    if any(m.lower() in head for m in ALIPAY_MARKERS):
        return "alipay"
    return "generic"


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _norm_key(value: Any) -> str:
    return re.sub(r"[\s_\-:/\\|（）()\[\]【】,，。.\u3000]+", "", _clean_text(value).lower())


def _parse_amount(val: Any) -> float:
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return round(abs(float(val)), 2)
    text = (
        _clean_text(val)
        .replace(",", "")
        .replace("¥", "")
        .replace("￥", "")
        .replace("元", "")
        .replace("CNY", "")
        .replace("RMB", "")
    )
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return 0.0
    try:
        return round(abs(float(match.group(0))), 2)
    except ValueError:
        return 0.0


def _raw_amount_negative(value: Any) -> bool:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value) < 0
    text = _clean_text(value).replace(",", "")
    return text.startswith("-") or (text.startswith("(") and text.endswith(")"))


DATE_KEYS = ("日期", "时间", "交易时间", "付款时间", "记账日期", "tx_date", "date", "time")
AMOUNT_KEYS = ("金额", "收支金额", "交易金额", "支出", "收入", "amount", "money", "fee")
TYPE_KEYS = ("类型", "收支", "收/支", "交易类型", "type", "direction")
CATEGORY_KEYS = ("分类", "账单分类", "类目", "category")
NOTE_KEYS = ("备注", "说明", "商品", "商品说明", "交易对方", "商户", "对方", "note", "memo", "description", "name")
ACCOUNT_KEYS = ("账户id", "账户ID", "account_id", "account")
SPLIT_KEYS = ("分摊", "split_type")
PAID_BY_KEYS = ("付款人", "paid_by")
ATTRIBUTED_TO_KEYS = ("归属人", "attributed_to")


def _find_key(row: dict[str, Any], aliases: tuple[str, ...]) -> str | None:
    normalized_aliases = [_norm_key(alias) for alias in aliases]
    for key in row:
        nkey = _norm_key(key)
        if any(alias and alias in nkey for alias in normalized_aliases):
            return key
    return None


def _field(row: dict[str, Any], aliases: tuple[str, ...], default: Any = "") -> Any:
    key = _find_key(row, aliases)
    return row.get(key, default) if key is not None else default


def _parse_date_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if 20000 <= float(value) <= 80000:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
    text = _clean_text(value)
    match = re.search(r"(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if match:
        y, m, d = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
        try:
            return date(y, m, d).isoformat()
        except ValueError:
            return local_today()
    compact = re.search(r"\b(\d{4})(\d{2})(\d{2})\b", text)
    if compact:
        y, m, d = (int(compact.group(1)), int(compact.group(2)), int(compact.group(3)))
        try:
            return date(y, m, d).isoformat()
        except ValueError:
            return local_today()
    return local_today()


def _parse_tx_type(type_raw: Any, amount_raw: Any = None) -> str:
    text = _clean_text(type_raw).lower()
    if any(word in text for word in ("income", "in", "收入", "收款", "入账", "退款")):
        return "income"
    if any(word in text for word in ("expense", "out", "支出", "付款", "消费", "扣款")):
        return "expense"
    return "expense" if _raw_amount_negative(amount_raw) else "expense"


def _default_category(tx_type: str) -> str:
    return "其他收入" if tx_type == "income" else "其他支出"


def _parse_generic_row(row: dict[str, Any]) -> dict[str, Any] | None:
    amount_key = _find_key(row, AMOUNT_KEYS)
    if not amount_key:
        return None
    amount_raw = row.get(amount_key)
    amount = _parse_amount(amount_raw)
    if amount <= 0:
        return None

    skip_text = " ".join(_clean_text(v) for v in row.values())
    if "不计" in skip_text or "忽略" in skip_text:
        return None

    tx_type = _parse_tx_type(_field(row, TYPE_KEYS, ""), amount_raw)
    return {
        "amount": amount,
        "category": _clean_text(_field(row, CATEGORY_KEYS, "")) or _default_category(tx_type),
        "type": tx_type,
        "note": _clean_text(_field(row, NOTE_KEYS, ""))[:100],
        "tx_date": _parse_date_value(_field(row, DATE_KEYS, local_today())),
        "account_id": _clean_text(_field(row, ACCOUNT_KEYS, "")) or None,
        "split_type": _clean_text(_field(row, SPLIT_KEYS, "none")) or "none",
        "paid_by": _clean_text(_field(row, PAID_BY_KEYS, "")) or None,
        "attributed_to": _clean_text(_field(row, ATTRIBUTED_TO_KEYS, "")) or None,
    }


def _parse_alipay_row(row: dict[str, str]) -> dict[str, Any] | None:
    time_key = next((k for k in row if "时间" in k), None)
    amount_key = next((k for k in row if "金额" in k), None)
    type_key = next((k for k in row if "收/支" in k or k == "收/支"), None)
    cat_key = next((k for k in row if "类型" in k or "分类" in k), None)
    note_key = next((k for k in row if "商品" in k or "备注" in k or "对方" in k), None)

    if not time_key or not amount_key:
        return None

    amount = _parse_amount(row.get(amount_key, "0"))
    if amount <= 0:
        return None

    tx_type_raw = row.get(type_key or "", "")
    tx_type = "income" if "收" in tx_type_raw or "收入" in tx_type_raw else "expense"
    if "不计" in tx_type_raw or "退款" in tx_type_raw:
        return None

    tx_date = _parse_date_value(row.get(time_key)) if row.get(time_key) else local_today()
    category = _clean_text(row.get(cat_key, "其他支出")) if cat_key else "其他支出"
    if tx_type == "income":
        category = "其他收入"

    return {
        "amount": amount,
        "category": category or "其他支出",
        "type": tx_type,
        "note": _clean_text(row.get(note_key))[:100],
        "tx_date": tx_date,
        "split_type": _clean_text(row.get("分摊") or row.get("split_type") or "none") or "none",
        "paid_by": _clean_text(row.get("付款人") or row.get("paid_by") or "") or None,
        "attributed_to": _clean_text(row.get("归属人") or row.get("attributed_to") or "") or None,
    }


def _parse_wechat_row(row: dict[str, str]) -> dict[str, Any] | None:
    time_key = next((k for k in row if "时间" in k), None)
    amount_key = next((k for k in row if "金额" in k), None)
    type_key = next((k for k in row if "收/支" in k), None)
    note_key = next((k for k in row if "商品" in k or "交易对方" in k), None)

    if not time_key or not amount_key:
        return None

    amount = _parse_amount(row.get(amount_key, "0"))
    if amount <= 0:
        return None

    tx_type_raw = row.get(type_key or "", "")
    tx_type = "income" if "/" in tx_type_raw and tx_type_raw.startswith("收") else "expense"
    if "收入" in tx_type_raw:
        tx_type = "income"

    tx_date = _parse_date_value(row.get(time_key))
    return {
        "amount": amount,
        "category": "其他收入" if tx_type == "income" else "其他支出",
        "type": tx_type,
        "note": _clean_text(row.get(note_key))[:100],
        "tx_date": tx_date,
    }


def _header_score(values: list[Any]) -> int:
    row = {str(index): value for index, value in enumerate(values)}
    score = 0
    if _find_key(row, DATE_KEYS):
        score += 2
    if _find_key(row, AMOUNT_KEYS):
        score += 2
    if _find_key(row, TYPE_KEYS):
        score += 1
    if _find_key(row, CATEGORY_KEYS):
        score += 1
    if _find_key(row, NOTE_KEYS):
        score += 1
    return score


def _find_header_index(rows: list[list[Any]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:30]):
        score = _header_score(row)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _dedupe_headers(headers: list[Any]) -> list[str]:
    result: list[str] = []
    seen: dict[str, int] = {}
    for index, header in enumerate(headers):
        text = _clean_text(header) or f"column_{index + 1}"
        count = seen.get(text, 0)
        seen[text] = count + 1
        result.append(text if count == 0 else f"{text}_{count + 1}")
    return result


def _dict_rows_from_matrix(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    start = _find_header_index(rows)
    headers = _dedupe_headers(rows[start])
    result: list[dict[str, Any]] = []
    for raw in rows[start + 1:]:
        if not any(_clean_text(value) for value in raw):
            continue
        padded = list(raw) + [""] * max(len(headers) - len(raw), 0)
        result.append({headers[index]: padded[index] for index in range(len(headers))})
    return result


def _parse_record_rows(fmt: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    parsed: list[dict[str, Any]] = []
    for row in rows:
        if fmt == "wechat":
            item = _parse_wechat_row(row) or _parse_generic_row(row)
        elif fmt == "alipay":
            item = _parse_alipay_row(row) or _parse_generic_row(row)
        else:
            item = _parse_generic_row(row) or _parse_alipay_row(row) or _parse_wechat_row(row)
        if item and item["amount"] > 0:
            parsed.append(item)
    return parsed


def _parse_csv_rows(content: str) -> tuple[str, list[dict[str, Any]]]:
    fmt = _detect_format(content)
    sample = "\n".join(content.splitlines()[:20])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    matrix = [row for row in csv.reader(io.StringIO(content), dialect)]
    rows = _dict_rows_from_matrix(matrix)
    return fmt, _parse_record_rows(fmt, rows)


def _parse_excel_rows(raw: bytes) -> tuple[str, list[dict[str, Any]]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Cannot read Excel file") from exc
    sheet = workbook.active
    matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
    head_text = "\n".join(" ".join(_clean_text(value) for value in row) for row in matrix[:20])
    detected = _detect_format(head_text)
    fmt = detected if detected != "generic" else "excel"
    rows = _dict_rows_from_matrix(matrix)
    return fmt, _parse_record_rows(detected, rows)


def _decode_csv(raw: bytes) -> str:
    for enc in ("utf-8-sig", "gbk", "gb18030", "utf-8"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    raise HTTPException(status_code=400, detail="Cannot decode CSV file")


def _parse_upload(file: UploadFile, raw: bytes) -> tuple[str, list[dict[str, Any]]]:
    filename = (file.filename or "").lower()
    content_type = (file.content_type or "").lower()
    if filename.endswith(".xlsx") or "spreadsheetml" in content_type:
        return _parse_excel_rows(raw)
    if filename.endswith(".xls"):
        raise HTTPException(status_code=400, detail="Please save .xls as .xlsx before importing")
    return _parse_csv_rows(_decode_csv(raw))


def _get_or_create_import_account(conn, user: dict, scope: str, couple_id: str | None) -> str:
    owner_col = "user_id" if scope == "personal" else "couple_id"
    owner_id = user["id"] if scope == "personal" else couple_id
    row = conn.execute(
        f"""
        SELECT id FROM accounts
        WHERE scope = ? AND {owner_col} = ? AND name IN ('账单导入', 'CSV导入') AND is_archived = 0
        LIMIT 1
        """,
        (scope, owner_id),
    ).fetchone()
    if row:
        return row["id"]
    account_id = new_id()
    now = utcnow()
    conn.execute(
        """
        INSERT INTO accounts (id, scope, user_id, couple_id, name, kind, balance, opening_balance, currency, is_archived, created_at)
        VALUES (?, ?, ?, ?, '账单导入', 'other', 0, 0, 'CNY', 0, ?)
        """,
        (account_id, scope, user["id"], couple_id, now),
    )
    return account_id


def _valid_account_id(conn, account_id: str | None, user: dict, scope: str, couple_id: str | None) -> str | None:
    if not account_id:
        return None
    owner_col = "user_id" if scope == "personal" else "couple_id"
    owner_id = user["id"] if scope == "personal" else couple_id
    row = conn.execute(
        f"""
        SELECT id FROM accounts
        WHERE id = ? AND scope = ? AND {owner_col} = ? AND is_archived = 0
        LIMIT 1
        """,
        (account_id, scope, owner_id),
    ).fetchone()
    return row["id"] if row else None


async def _commit_import_items(
    items: list[dict[str, Any]],
    scope: str,
    current_user: dict,
) -> dict[str, Any]:
    couple_id = current_user.get("couple_id") if scope == "couple" else None
    owner_col = "user_id" if scope == "personal" else "couple_id"
    owner_id = current_user["id"] if scope == "personal" else couple_id
    now = utcnow()
    imported = 0
    skipped = 0
    batch_seen: set[tuple] = set()
    with get_db() as conn:
        import_account_id = _get_or_create_import_account(conn, current_user, scope, couple_id)
        for item in items:
            key = (item["tx_date"], item["type"], round(item["amount"], 2), (item.get("note") or "")[:80])
            if key in batch_seen:
                skipped += 1
                continue
            batch_seen.add(key)
            exists = conn.execute(
                f"""
                SELECT id FROM transactions
                WHERE scope = ? AND {owner_col} = ?
                  AND tx_date = ? AND type = ? AND abs(amount - ?) < 0.001
                  AND COALESCE(note, '') = ?
                LIMIT 1
                """,
                (scope, owner_id, item["tx_date"], item["type"], item["amount"], item.get("note") or ""),
            ).fetchone()
            if exists:
                skipped += 1
                continue
            tx_id = new_id()
            paid_by = item.get("paid_by") or current_user["id"]
            split_type = item.get("split_type") or "none"
            attributed_to = item.get("attributed_to")
            target_account_id = _valid_account_id(conn, item.get("account_id"), current_user, scope, couple_id) or import_account_id
            paid_by, attributed_to, split_type = _validate_split_fields(
                conn, current_user, scope, couple_id, paid_by, attributed_to, split_type,
            )
            conn.execute(
                """
                INSERT INTO transactions
                (id, scope, user_id, couple_id, amount, category, type, note, tx_date,
                 created_by, created_at, account_id, tx_kind, split_type, paid_by, attributed_to)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'normal', ?, ?, ?)
                """,
                (
                    tx_id, scope, current_user["id"], couple_id,
                    item["amount"], item["category"], item["type"],
                    item["note"], item["tx_date"], current_user["id"], now, target_account_id,
                    split_type, paid_by, attributed_to,
                ),
            )
            delta = item["amount"] if item["type"] == "income" else -item["amount"]
            apply_balance_delta(conn, target_account_id, delta)
            imported += 1

    if scope == "couple" and couple_id and imported:
        await manager.broadcast_couple(
            couple_id,
            {"event": "transactions_imported", "data": {"count": imported, "scope": scope}},
        )

    return {"imported": imported, "skipped": skipped, "account_id": import_account_id}


async def _import_upload(
    file: UploadFile,
    scope: Literal["personal", "couple"],
    dry_run: bool,
    current_user: dict,
) -> dict[str, Any]:
    _validate_scope_access(current_user, scope)
    raw = await file.read()
    if len(raw) > 15 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Import file too large (max 15MB)")
    fmt, items = _parse_upload(file, raw)
    if not items:
        raise HTTPException(status_code=400, detail="No valid transactions found in import file")

    base = {
        "format": fmt,
        "filename": file.filename or "",
        "count": len(items),
        "preview": items[:10],
    }
    if dry_run:
        return {**base, "imported": 0, "skipped": 0}

    result = await _commit_import_items(items, scope, current_user)
    return {
        **base,
        **result,
    }


@router.post("/import/csv")
async def import_csv(
    file: UploadFile = File(...),
    scope: Literal["personal", "couple"] = Query("personal"),
    dry_run: bool = Query(False),
    current_user: dict = Depends(get_current_user),
) -> dict[str, Any]:
    return await _import_upload(file, scope, dry_run, current_user)


@router.post("/import/excel")
async def import_excel(
    file: UploadFile = File(...),
    scope: Literal["personal", "couple"] = Query("personal"),
    dry_run: bool = Query(False),
    current_user: dict = Depends(get_current_user),
) -> dict[str, Any]:
    return await _import_upload(file, scope, dry_run, current_user)


@router.post("/import/bills")
async def import_bills(
    file: UploadFile = File(...),
    scope: Literal["personal", "couple"] = Query("personal"),
    dry_run: bool = Query(False),
    current_user: dict = Depends(get_current_user),
) -> dict[str, Any]:
    return await _import_upload(file, scope, dry_run, current_user)


@router.get("/summary")
def data_summary(
    scope: Literal["personal", "couple"] = Query("personal"),
    current_user: dict = Depends(get_current_user),
) -> dict[str, Any]:
    """Account ledger stats for profile / data overview."""
    _validate_scope_access(current_user, scope)

    if scope == "personal":
        tx_where = "scope = 'personal' AND user_id = ?"
        tx_param = current_user["id"]
        owner_col = "user_id"
        owner_id = current_user["id"]
    else:
        tx_where = "scope = 'couple' AND couple_id = ?"
        tx_param = current_user["couple_id"]
        owner_col = "couple_id"
        owner_id = current_user["couple_id"]

    with get_db() as conn:
        tx_row = conn.execute(
            f"SELECT COUNT(*) AS cnt, MIN(tx_date) AS oldest, MAX(tx_date) AS newest "
            f"FROM transactions WHERE {tx_where}",
            (tx_param,),
        ).fetchone()
        acc_count = conn.execute(
            f"SELECT COUNT(*) AS cnt FROM accounts "
            f"WHERE scope = ? AND {owner_col} = ? AND is_archived = 0",
            (scope, owner_id),
        ).fetchone()["cnt"]

    return {
        "scope": scope,
        "tx_count": int(tx_row["cnt"] or 0),
        "oldest_tx_date": tx_row["oldest"][:10] if tx_row["oldest"] else None,
        "newest_tx_date": tx_row["newest"][:10] if tx_row["newest"] else None,
        "account_count": int(acc_count or 0),
    }


@router.get("/export/csv")
def export_csv(
    scope: Literal["personal", "couple"] = Query("personal"),
    month: str | None = Query(None),
    current_user: dict = Depends(get_current_user),
) -> StreamingResponse:
    _validate_scope_access(current_user, scope)

    query = "SELECT tx_date, category, type, amount, note, paid_by, split_type, attributed_to FROM transactions WHERE scope = ?"
    params: list[Any] = [scope]
    if scope == "personal":
        query += " AND user_id = ?"
        params.append(current_user["id"])
    else:
        query += " AND couple_id = ?"
        params.append(current_user["couple_id"])
    if month:
        query += " AND tx_date LIKE ?"
        params.append(f"{month}%")
    query += " ORDER BY tx_date DESC"

    with get_db() as conn:
        rows = conn.execute(query, params).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["日期", "分类", "类型", "金额", "备注", "付款人", "分摊", "归属人"])
    for r in rows:
        writer.writerow([
            r["tx_date"][:10], r["category"], r["type"], r["amount"], r["note"] or "",
            r["paid_by"] or "", r["split_type"] or "none", r["attributed_to"] or "",
        ])

    output.seek(0)
    filename = f"ledger_{scope}_{month or 'all'}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/excel")
def export_excel(
    scope: Literal["personal", "couple"] = Query("personal"),
    month: str | None = Query(None),
    current_user: dict = Depends(get_current_user),
) -> Response:
    from openpyxl import Workbook

    _validate_scope_access(current_user, scope)

    query = "SELECT tx_date, category, type, amount, note, account_id, paid_by, split_type, attributed_to FROM transactions WHERE scope = ?"
    params: list[Any] = [scope]
    if scope == "personal":
        query += " AND user_id = ?"
        params.append(current_user["id"])
    else:
        query += " AND couple_id = ?"
        params.append(current_user["couple_id"])
    if month:
        query += " AND tx_date LIKE ?"
        params.append(f"{month}%")
    query += " ORDER BY tx_date DESC"

    with get_db() as conn:
        rows = conn.execute(query, params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "账目"
    ws.append(["日期", "分类", "类型", "金额", "备注", "账户ID", "付款人", "分摊", "归属人"])
    for r in rows:
        ws.append([
            r["tx_date"][:10], r["category"], r["type"], r["amount"], r["note"] or "", r["account_id"] or "",
            r["paid_by"] or "", r["split_type"] or "none", r["attributed_to"] or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"ledger_{scope}_{month or 'all'}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
